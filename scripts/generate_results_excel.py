"""Generate a pre-filled Excel results template for the tournament admin.

Creates data/results_template.xlsx with two sheets:
  - Group Stage: all 48 teams sorted by group, with group-stage stat columns
  - Knockout Rounds: all 48 teams with knockout stats and Round Reached

Fill this in during the tournament, then import via Admin -> Results Entry -> Upload Excel.

Usage:
    .\.venv\Scripts\python.exe scripts/generate_results_excel.py
"""
import sys
from pathlib import Path

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


TIER_COLORS = {1: "105AAC", 2: "15803D", 3: "A16207", 4: "B91C1C"}
HEADER_BG   = "0D1B2A"
HEADER_FG   = "D4A017"

THIN = Border(
    left=Side(style="thin", color="444444"),
    right=Side(style="thin", color="444444"),
    top=Side(style="thin", color="444444"),
    bottom=Side(style="thin", color="444444"),
)


def _header_cell(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = PatternFill("solid", fgColor=HEADER_BG)
    c.font = Font(color=HEADER_FG, bold=True)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = THIN
    return c


def _data_cell(ws, row, col, value, tier=None, number=True):
    c = ws.cell(row=row, column=col, value=value)
    if tier and tier in TIER_COLORS:
        c.fill = PatternFill("solid", fgColor=TIER_COLORS[tier])
        c.font = Font(color="FFFFFF")
    c.alignment = Alignment(horizontal="center" if number else "left", vertical="center")
    c.border = THIN
    return c


def build_group_stage_sheet(ws, teams):
    ws.title = "Group Stage"
    ws.freeze_panes = "A2"

    headers = [
        "Team", "Group", "Tier",
        "Goals", "Clean Sheets", "Penalty Wins", "Comeback Wins", "Group Winner (1=Yes)",
    ]
    col_widths = [22, 8, 6, 8, 14, 14, 16, 20]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        _header_cell(ws, 1, col, h)
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 25

    sorted_teams = sorted(teams, key=lambda r: (r["Group"], r["Team"]))
    for i, team in enumerate(sorted_teams, 2):
        tier = int(team["Tier"])
        _data_cell(ws, i, 1, team["Team"], tier=tier, number=False)
        _data_cell(ws, i, 2, team["Group"], number=False)
        _data_cell(ws, i, 3, tier)
        for col in range(4, 9):
            _data_cell(ws, i, col, 0)

    ws.add_table = None


def build_knockout_sheet(ws, teams):
    ws.title = "Knockout Rounds"
    ws.freeze_panes = "A2"

    headers = [
        "Team", "Group", "Tier",
        "KO Goals", "KO Clean Sheets", "KO Penalty Wins", "KO Comeback Wins",
        "Round Reached",
    ]
    col_widths = [22, 8, 6, 10, 16, 16, 18, 16]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        _header_cell(ws, 1, col, h)
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 25

    # Dropdown validation for Round Reached column (H = col 8)
    dv = DataValidation(
        type="list",
        formula1='"GroupStage,R16,QF,SF,Final,Winner"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.sqref = f"H2:H{len(teams) + 1}"
    ws.add_data_validation(dv)

    sorted_teams = sorted(teams, key=lambda r: (r["Group"], r["Team"]))
    for i, team in enumerate(sorted_teams, 2):
        tier = int(team["Tier"])
        _data_cell(ws, i, 1, team["Team"], tier=tier, number=False)
        _data_cell(ws, i, 2, team["Group"], number=False)
        _data_cell(ws, i, 3, tier)
        for col in range(4, 8):
            _data_cell(ws, i, col, 0)
        c = ws.cell(row=i, column=8, value="")
        c.alignment = Alignment(horizontal="center")
        c.border = THIN


def build_instructions_sheet(ws):
    ws.title = "Instructions"
    ws.sheet_view.showGridLines = False

    lines = [
        ("WC 2026 Sweepstake — Results Template", True),
        ("", False),
        ("HOW TO USE THIS FILE", True),
        ("", False),
        ("1. Fill in the GROUP STAGE sheet after all group games are played.", False),
        ("   - Enter cumulative totals for each team.", False),
        ("   - Group Winner: put 1 if the team finished top of their group, 0 otherwise.", False),
        ("", False),
        ("2. Fill in the KNOCKOUT ROUNDS sheet as teams are eliminated.", False),
        ("   - KO Goals / Clean Sheets / Penalty Wins / Comeback Wins: knockout matches only.", False),
        ("   - Round Reached: the furthest round a team reached.", False),
        ("     GroupStage = eliminated in groups", False),
        ("     R16 = reached the Round of 16 but went out there", False),
        ("     QF  = reached the Quarter Final but went out", False),
        ("     SF  = reached the Semi Final but went out", False),
        ("     Final = reached the Final but lost", False),
        ("     Winner = won the World Cup", False),
        ("", False),
        ("3. Save this file as results_template.xlsx in the data/ folder.", False),
        ("", False),
        ("4. Go to Admin -> Results Entry -> Upload Excel to import into the app.", False),
        ("", False),
        ("WHAT ROUND REACHED MEANS", True),
        ("", False),
        ("The Round Reached field tells the app how far each team progressed.", False),
        ("Teams that are still active (not yet eliminated) should be left blank.", False),
        ("Update this field each time a team is knocked out.", False),
    ]

    for row, (text, bold) in enumerate(lines, 1):
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(bold=bold, size=11 if bold else 10)
        c.alignment = Alignment(vertical="center")

    ws.column_dimensions["A"].width = 70


def main():
    from src.team_database import load_teams
    teams_df = load_teams()
    teams = teams_df[["Team", "Group", "Tier"]].to_dict("records")

    wb = Workbook()
    wb.remove(wb.active)

    ws_gs = wb.create_sheet("Group Stage")
    ws_ko = wb.create_sheet("Knockout Rounds")
    ws_in = wb.create_sheet("Instructions")

    build_group_stage_sheet(ws_gs, teams)
    build_knockout_sheet(ws_ko, teams)
    build_instructions_sheet(ws_in)

    out = ROOT / "data" / "results_template.xlsx"
    wb.save(out)
    print(f"Template saved: {out}")
    print(f"  {len(teams)} teams across {len({t['Group'] for t in teams})} groups")


if __name__ == "__main__":
    main()
